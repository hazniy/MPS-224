#Task 1
import openpyxl
import os

def create_spreadsheet(filename, strings):
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "String"
    ws['B1'] = "Length"

    for i, s in enumerate(strings, start=2):
        ws[f'A{i}'] = s
        ws[f'B{i}'] = len(s)

    total_row = len(strings) + 2
    ws[f'A{total_row}'] = "Total"
    ws[f'B{total_row}'] = f"=SUM(B2:B{total_row - 1})"

    wb.save(filename)

    # This line tells you exactly where it is
    print("Saved to:", os.path.abspath(filename))
    
create_spreadsheet(
    'cheese.xlsx',
    ['Cheddar', 'Wensleydale', 'Stilton']
)

#submit
def create_spreadsheet(filename, strings):
    """
    Creates and saves an Excel spreadsheet containing a list of strings and their lengths.

    The spreadsheet will have two columns: "String" and "Length". Each row contains
    a string from the input list and its corresponding length. A final row is added
    with the label "Total" and a formula that calculates the sum of all string lengths.

    Parameters:
        filename (str): The name of the Excel file to be created (must end with .xlsx).
        strings (list of str): A list of strings to include in the spreadsheet.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "String"
    ws['B1'] = "Length"
    for i, s in enumerate(strings, start=2):
        ws[f'A{i}'] = s
        ws[f'B{i}'] = len(s)
    total_row = len(strings) + 2
    ws[f'A{total_row}'] = "Total"
    ws[f'B{total_row}'] = f"=SUM(B2:B{total_row - 1})"
    wb.save(filename)

#Task 2
from PIL import Image
from IPython.display import display

# Open the original image
img = Image.open("python_facing_left.jpg")

# Create mirror image (flip horizontally)
mirrored = img.transpose(Image.FLIP_LEFT_RIGHT)

# Save the mirrored image
mirrored.save("python_facing_right.jpg")

# Display both images
print("Original image:")
display(img)

print("Mirrored image:")
display(mirrored)

#Task 3
#get the data from hugging face
import requests
from IPython.display import display

def get_data():
    url = "https://datasets-server.huggingface.co/rows?dataset=ceyda/smithsonian_butterflies&config=default&split=train&offset=0&length=100"
    
    dataset = requests.get(url).json()
    data = dataset['rows']
    
    return data

#extract and show a butterfly image 
from PIL import Image

def show_butterfly(data, k):
    # Extract the actual URL string
    image_url = data[k]['row']['image']['src']
    
    # Load and display image
    img = Image.open(requests.get(image_url, stream=True).raw)
    display(img)

#show a random butterfly 
import random

def show_random_butterfly(data):
    k = random.randint(0, len(data) - 1)
    show_butterfly(data, k)
    
#extract the butterfly family 
def family(data, k):
    taxonomy = data[k]['row']['taxonomy']
    
    # Split and clean
    parts = [x.strip() for x in taxonomy.split(',')]
    
    return parts[5]

data = get_data()

# Show a specific butterfly
show_butterfly(data, 42)

# Show a random one
show_random_butterfly(data)

# Get its family
print(family(data, 0))
